from __future__ import annotations

import sys
import threading
import time
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
_A_APP = _ROOT / "apps" / "auto-extract"
_A_SRC = _A_APP / "src"
_IM_SRC = _ROOT / "apps" / "im-module" / "src"
for p in (_A_SRC, _A_APP, _IM_SRC, _ROOT):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))


def test_task_store_persist_and_update(tmp_path, monkeypatch):
    for p in list(sys.path):
        if "im-module" in p.replace("\\", "/"):
            sys.path.remove(p)
    if str(_A_SRC) not in sys.path:
        sys.path.insert(0, str(_A_SRC))
    for name in ("config", "task_store", "models"):
        sys.modules.pop(name, None)

    import config
    import task_store

    db = tmp_path / "tasks.db"
    monkeypatch.setattr(config, "TASKS_DB", db)
    task_store._conn = None
    task_store.open_store()

    from models import Task

    t = Task(
        task_id="t-0001",
        url="https://x/a.apk",
        source_file="im_1.json",
        im_chat_id="group:cid-a",
    )
    task_store.insert_task(t)
    got = task_store.get_task("t-0001")
    assert got is not None and got.status == "queued"
    assert got.im_chat_id == "group:cid-a"

    updated = task_store.update_task("t-0001", status="downloaded", filename="a.apk")
    assert updated is not None and updated.status == "downloaded"
    assert updated.filename == "a.apk"
    assert updated.im_chat_id == "group:cid-a"

    again = task_store.get_task("t-0001")
    assert again is not None and again.status == "downloaded"


def test_recovery_matrix_routes(tmp_path, monkeypatch):
    # Ensure auto-extract config wins over im-module config on sys.path.
    for p in list(sys.path):
        if p.replace("\\", "/").endswith("im-module/src"):
            sys.path.remove(p)
    if str(_A_SRC) not in sys.path:
        sys.path.insert(0, str(_A_SRC))

    import importlib

    for name in (
        "config",
        "task_store",
        "queue_manager",
        "pipeline",
        "pipeline_queues",
        "models",
    ):
        sys.modules.pop(name, None)

    import config
    import pipeline
    import pipeline_queues as pq
    import task_store
    from models import Task

    db = tmp_path / "tasks.db"
    monkeypatch.setattr(config, "TASKS_DB", db)
    monkeypatch.setattr(config, "DOWNLOADS_DIR", tmp_path / "dl")
    config.DOWNLOADS_DIR.mkdir()
    task_store._conn = None
    task_store.open_store()

    for q in (pq.Q_DOWNLOAD, pq.Q_PATCH, pq.Q_DEVICE, pq.Q_EXTRACT, pq.Q_ARCHIVE):
        while not q.empty():
            try:
                q.get_nowait()
                q.task_done()
            except Exception:
                break

    task_store.insert_task(
        Task(task_id="t-q", url="https://x/q.apk", source_file="a.json", status="queued")
    )
    apk = config.DOWNLOADS_DIR / "d.apk"
    apk.write_bytes(b"apk")
    task_store.insert_task(
        Task(
            task_id="t-d",
            url="https://x/d.apk",
            source_file="a.json",
            status="downloaded",
            filename="d.apk",
        )
    )
    task_store.insert_task(
        Task(
            task_id="t-od",
            url="https://x/od.apk",
            source_file="a.json",
            status="on_device",
        )
    )
    task_store.insert_task(
        Task(
            task_id="t-oe",
            url="https://x/oe.apk",
            source_file="a.json",
            status="on_extract",
        )
    )

    pipeline.recover_and_enqueue()
    time.sleep(0.05)

    assert task_store.get_task("t-od").status == "failed"
    assert task_store.get_task("t-oe").status == "failed"


def test_adb_pool_exclusive(monkeypatch):
    from resource_pools import AdbPool

    pool = AdbPool(wait_timeout_sec=1.0)
    monkeypatch.setattr(
        pool,
        "refresh",
        lambda: ["dev-a"],
    )
    # inject fake lock map
    import threading

    with pool._cond:
        pool._locks = {"dev-a": threading.Lock()}

    held = []

    def hold():
        s = pool.acquire()
        held.append(s)
        time.sleep(0.3)
        pool.release(s)

    t1 = threading.Thread(target=hold)
    t2 = threading.Thread(target=hold)
    t1.start()
    t2.start()
    t1.join()
    t2.join()
    assert held == ["dev-a", "dev-a"]


def test_filename_overwrite_bumps_run_gen(tmp_path, monkeypatch):
    for p in list(sys.path):
        if "im-module" in p.replace("\\", "/"):
            sys.path.remove(p)
    if str(_A_SRC) not in sys.path:
        sys.path.insert(0, str(_A_SRC))
    for name in (
        "config",
        "task_store",
        "queue_manager",
        "pipeline_queues",
        "models",
        "downloader",
    ):
        sys.modules.pop(name, None)

    import config
    import pipeline_queues as pq
    import queue_manager
    import task_store

    db = tmp_path / "tasks.db"
    state = tmp_path / "state"
    state.mkdir()
    monkeypatch.setattr(config, "TASKS_DB", db)
    monkeypatch.setattr(config, "STATE_DIR", state)
    monkeypatch.setattr(config, "QUEUE_STATUS_FILE", state / "q.json")
    task_store._conn = None
    queue_manager.load()

    # Drain any recover leftovers.
    while not pq.Q_DOWNLOAD.empty():
        try:
            pq.Q_DOWNLOAD.get_nowait()
            pq.Q_DOWNLOAD.task_done()
        except Exception:
            break

    first = queue_manager.enqueue_urls(
        ["https://cdn.example/game.apk"],
        "im_a.json",
        im_chat_id="group:old",
        im_sender_id="u1",
    )
    assert len(first) == 1
    tid = first[0].task_id
    assert first[0].filename == "game.apk"
    assert first[0].run_gen == 0

    # Simulate finished run with delivery markers.
    done = queue_manager.update_task(
        tid,
        status="success",
        label="OldGame",
        result_csv=str(tmp_path / "old.csv"),
        buf_done_zip=str(tmp_path / "old.zip"),
        session_id="ses-old",
        error="",
        im_delivered_at="2026-01-01T00:00:00Z",
        finished_at="2026-01-01T00:00:00Z",
        expected_run_gen=0,
    )
    assert done is not None and done.status == "success"

    second = queue_manager.enqueue_urls(
        ["https://cdn.example/game.apk?x=1"],
        "im_b.json",
        im_chat_id="group:new",
        im_sender_id="u2",
    )
    assert len(second) == 1
    assert second[0].task_id == tid
    assert second[0].run_gen == 1
    assert second[0].status == "queued"
    assert second[0].filename == "game.apk"
    assert second[0].label == ""
    assert second[0].result_csv == ""
    assert second[0].buf_done_zip == ""
    assert second[0].session_id == ""
    assert second[0].im_delivered_at == ""
    assert second[0].finished_at == ""
    assert second[0].im_chat_id == "group:new"
    assert second[0].source_file == "im_b.json"

    # Stale gen must not rewrite ledger.
    stale = queue_manager.update_task(
        tid, status="failed", error="stale", expected_run_gen=0
    )
    assert stale is None
    cur = queue_manager.get_task(tid)
    assert cur is not None
    assert cur.status == "queued"
    assert cur.run_gen == 1
    assert cur.error == ""

    # Unique index: only one row for this filename.
    import sqlite3

    conn = sqlite3.connect(str(db))
    n = conn.execute(
        "SELECT COUNT(*) FROM tasks WHERE filename=?", ("game.apk",)
    ).fetchone()[0]
    conn.close()
    assert n == 1


def test_filename_dedupe_on_migrate(tmp_path, monkeypatch):
    for p in list(sys.path):
        if "im-module" in p.replace("\\", "/"):
            sys.path.remove(p)
    if str(_A_SRC) not in sys.path:
        sys.path.insert(0, str(_A_SRC))
    for name in ("config", "task_store", "models"):
        sys.modules.pop(name, None)

    import sqlite3

    import config
    import task_store

    db = tmp_path / "legacy.db"
    monkeypatch.setattr(config, "TASKS_DB", db)
    conn = sqlite3.connect(str(db))
    conn.execute(
        """
        CREATE TABLE tasks (
            task_id TEXT PRIMARY KEY,
            url TEXT NOT NULL,
            source_file TEXT NOT NULL DEFAULT '',
            filename TEXT NOT NULL DEFAULT '',
            label TEXT NOT NULL DEFAULT '',
            labels_json TEXT NOT NULL DEFAULT '{}',
            status TEXT NOT NULL,
            error TEXT NOT NULL DEFAULT '',
            result_csv TEXT NOT NULL DEFAULT '',
            session_id TEXT NOT NULL DEFAULT '',
            buf_done_zip TEXT NOT NULL DEFAULT '',
            adb_serial TEXT NOT NULL DEFAULT '',
            hotfix_has_files TEXT NOT NULL DEFAULT '',
            hotfix_pull_source TEXT NOT NULL DEFAULT '',
            screen_reached TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            finished_at TEXT NOT NULL DEFAULT '',
            im_delivered_at TEXT NOT NULL DEFAULT '',
            im_chat_id TEXT NOT NULL DEFAULT '',
            im_sender_id TEXT NOT NULL DEFAULT '',
            im_deliver_error TEXT NOT NULL DEFAULT ''
        )
        """
    )
    conn.execute(
        "INSERT INTO tasks(task_id,url,filename,status,created_at,updated_at) "
        "VALUES (?,?,?,?,?,?)",
        ("t-old", "https://x/a.apk", "a.apk", "failed", "t1", "2026-01-01T00:00:00Z"),
    )
    conn.execute(
        "INSERT INTO tasks(task_id,url,filename,status,created_at,updated_at) "
        "VALUES (?,?,?,?,?,?)",
        ("t-new", "https://x/a.apk", "a.apk", "success", "t2", "2026-02-01T00:00:00Z"),
    )
    conn.commit()
    conn.close()

    task_store._conn = None
    task_store.open_store()
    assert task_store.get_task("t-old") is None
    kept = task_store.get_task("t-new")
    assert kept is not None and kept.status == "success"
    assert task_store.get_task_by_filename("a.apk").task_id == "t-new"


def test_ops_commands_parse():
    # Prefer im-module ops_commands
    if str(_A_SRC) in sys.path:
        sys.path.remove(str(_A_SRC))
    sys.path.insert(0, str(_IM_SRC))
    import importlib

    import ops_commands

    importlib.reload(ops_commands)
    assert ops_commands.parse_ops_command("query progress").kind == "query_progress"
    assert ops_commands.parse_ops_command("query export").kind == "export_usage"
    assert ops_commands.parse_ops_command("export aaa").kind == "export_usage"
    assert ops_commands.parse_ops_command("export table all").kind == "export_table"
    assert ops_commands.parse_ops_command("export table all").arg == "all"
    assert ops_commands.parse_ops_command("export table success").arg == "success"
    assert ops_commands.parse_ops_command("export").kind == "help"
    assert ops_commands.parse_ops_command("query password").kind == "query_password"
    assert ops_commands.parse_ops_command("ping").kind == "ping"
    assert ops_commands.parse_ops_command("PING!").kind == "ping"
    assert ops_commands.parse_ops_command("query label 三国").kind == "query_label"
    assert ops_commands.parse_ops_command("query label 三国").arg == "三国"
    assert ops_commands.parse_ops_command("query gid t-1").kind == "query_gid"
    assert ops_commands.parse_ops_command("query gid t-1").arg == "t-1"
    assert ops_commands.parse_ops_command("query status success").arg == "success"
    assert ops_commands.parse_ops_command("query foo").kind == "query_usage"
    assert ops_commands.parse_ops_command("help").kind == "help"
    assert ops_commands.parse_ops_command("你好").kind == "greet"
    assert ops_commands.parse_ops_command("你好呀！").kind == "greet"
    assert ops_commands.parse_ops_command("Hello").kind == "greet"
    assert ops_commands.parse_ops_command("query all").kind == "query_usage"

    label, statuses = ops_commands.resolve_status_filter("running")
    assert label == "进行中"
    assert statuses == ops_commands.ACTIVE_STATUSES
    label, statuses = ops_commands.resolve_status_filter("failed")
    assert label == "失败"
    assert "success" not in statuses
    assert "timeout" in statuses
    assert ops_commands.resolve_status_filter("失败") == (
        "失败",
        frozenset(s for s in ops_commands.TERMINAL_STATUSES if s != "success"),
    )
    assert ops_commands.resolve_status_filter("queued") == (
        "排队中",
        frozenset({"queued"}),
    )
    assert ops_commands.resolve_status_filter("nope") is None
    assert ops_commands.status_label_zh("on_device") == "设备处理中"
    assert ops_commands.status_label_zh("success") == "成功"
    assert ops_commands.parse_ops_command("query top_n 20").kind == "query_usage"
    assert ops_commands.parse_ops_command("查询表格 all") is None
    assert ops_commands.parse_ops_command('{"get-texts":{"urls":["x"]}}') is None


def test_ledger_query_text(tmp_path, monkeypatch):
    if str(_A_SRC) in sys.path:
        sys.path.remove(str(_A_SRC))
    sys.path.insert(0, str(_IM_SRC))
    import importlib
    import sqlite3

    import config as im_config

    importlib.reload(im_config)
    db = tmp_path / "tasks.db"
    export = tmp_path / "exports"
    monkeypatch.setattr(im_config, "TASKS_DB", db)
    monkeypatch.setattr(im_config, "QUERY_EXPORT_DIR", export)
    monkeypatch.setattr(im_config, "ZIP_PASSWORD", "gametool999")

    conn = sqlite3.connect(str(db))
    conn.execute(
        """
        CREATE TABLE tasks (
            task_id TEXT PRIMARY KEY,
            url TEXT, label TEXT, filename TEXT, status TEXT, error TEXT,
            result_csv TEXT, session_id TEXT, buf_done_zip TEXT, source_file TEXT,
            adb_serial TEXT, created_at TEXT, updated_at TEXT, finished_at TEXT,
            im_delivered_at TEXT, im_chat_id TEXT
        )
        """
    )
    conn.execute(
        "INSERT INTO tasks VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (
            "t-0001",
            "https://x/a.apk",
            "Game",
            "a.apk",
            "success",
            "",
            "",
            "ses-1",
            str(tmp_path / "a.zip"),
            "im_1.json",
            "emulator-5554",
            "2026-08-04T12:01:00Z",
            "2026-08-04T12:01:00Z",
            "2026-08-04T12:01:00Z",
            "2026-08-04T12:05:00Z",
            "group:cid-a",
        ),
    )
    (tmp_path / "a.zip").write_bytes(b"x")
    conn.execute(
        "INSERT INTO tasks VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (
            "t-0002",
            "https://x/b.apk",
            "ActiveGame",
            "b.apk",
            "on_extract",
            "",
            "",
            "",
            "",
            "im_1.json",
            "",
            "2026-08-04T12:00:00Z",
            "2026-08-04T12:00:30Z",
            "",
            "",
            "group:cid-a",
        ),
    )
    conn.execute(
        "INSERT INTO tasks VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (
            "t-0003",
            "https://x/a-old.apk",
            "GameOld",
            "a.apk",
            "failed",
            "old",
            "",
            "",
            "",
            "im_0.json",
            "",
            "2026-08-03T12:00:00Z",
            "2026-08-03T12:00:00Z",
            "2026-08-03T12:00:00Z",
            "",
            "group:cid-a",
        ),
    )
    conn.commit()
    conn.close()

    import ops_commands
    import task_ledger_query

    importlib.reload(ops_commands)
    importlib.reload(task_ledger_query)

    assert task_ledger_query.to_shanghai("2026-08-04T12:01:00Z") == "2026-08-04 20:01:00"

    pw = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="query_password")
    )
    assert pw.ok and pw.message == (
        "解压密码：gametool999\n说明：群内回传的是加密 .zip，用解压工具输入密码即可。"
    )

    prog = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="query_progress")
    )
    assert prog.ok and "全部进行中" in prog.message and "t-0002" in prog.message
    assert "t-0001" not in prog.message
    assert prog.file_path is None

    gid = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="query_gid", arg="t-0001")
    )
    assert gid.ok and "2026-08-04 20:01:00" in gid.message
    assert "任务号：t-0001" in gid.message
    assert "群回传：2026-08-04 20:05:00" in gid.message
    assert "结果 zip：已生成" in gid.message
    assert "session" not in gid.message
    assert "adb" not in gid.message
    assert "emulator-5554" not in gid.message
    assert "ses-1" not in gid.message

    exported = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="export_table", arg="all")
    )
    assert exported.ok and exported.file_path and exported.file_path.is_file()
    assert exported.file_path.suffix == ".xlsx"
    # a.apk appears twice in DB; unique keeps latest (t-0001), plus b.apk → 2
    assert exported.row_count == 2
    assert "已导出表格" in exported.message
    assert "全部" in exported.message
    assert "按文件名取最新" in exported.message
    from openpyxl import load_workbook

    wb = load_workbook(exported.file_path, read_only=True)
    sheet_rows = list(wb.active.iter_rows(values_only=True))
    assert sheet_rows[0] == (
        "filename",
        "label",
        "status",
        "error",
        "updated_at",
        "finished_at",
    )
    body = sheet_rows[1:]
    assert any(r[0] == "a.apk" and r[1] == "Game" for r in body)
    assert not any(r[0] == "a.apk" and r[1] == "GameOld" for r in body)
    a_row = next(r for r in body if r[0] == "a.apk")
    assert a_row[2] == "成功"
    assert a_row[3] in ("", None)
    assert a_row[4] == "2026-08-04: 20:01"
    assert a_row[5] == "2026-08-04: 20:01"

    success_only = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="export_table", arg="success")
    )
    assert success_only.ok and success_only.row_count == 1
    assert "成功" in success_only.message

    running_only = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="export_table", arg="running")
    )
    assert running_only.ok and running_only.row_count == 1
    assert "进行中" in running_only.message

    # Latest a.apk is success → failed range excludes it after unique-first
    failed_only = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="export_table", arg="failed")
    )
    assert failed_only.ok and failed_only.row_count == 0
    assert "失败" in failed_only.message

    cn_failed = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="export_table", arg="失败")
    )
    assert cn_failed.ok and cn_failed.row_count == 0

    empty_status = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="query_status", arg="")
    )
    assert not empty_status.ok
    assert "query status" in empty_status.message
    assert "running" in empty_status.message

    gid_zh = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="query_gid", arg="t-0001")
    )
    assert gid_zh.ok and "状态：成功" in gid_zh.message

    bad_shape = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="export_usage")
    )
    assert not bad_shape.ok
    assert "export table" in bad_shape.message
    assert "范围" in bad_shape.message
    assert "【提交任务】" not in bad_shape.message

    bad = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="export_table", arg="nope")
    )
    assert not bad.ok
    assert "范围不正确" in bad.message
    assert "all" in bad.message
    assert "running" in bad.message
    assert "success" in bad.message
    assert "failed" in bad.message
    assert "【提交任务】" not in bad.message

    miss = task_ledger_query.run_ledger_query(
        ops_commands.OpsCommand(kind="query_gid", arg="nope")
    )
    assert miss.ok and miss.row_count == 0 and "未找到" in miss.message